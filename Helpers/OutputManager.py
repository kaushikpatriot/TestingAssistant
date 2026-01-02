from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

class ExcelManager:
    def __init__(self, mode='new', filepath=None):
        self.filepath = filepath
        if mode == 'new':
            self.wb = Workbook()
            self.filepath = filepath
        else:
            if filepath:
                self.wb = load_workbook(filepath)
                self.sheetnames = self.wb.sheetnames
            else:
                raise Exception("Filepath required to load an existing workbook")
    
    def createWorksheet(self, sheetName):
        self.wb.create_sheet(sheetName)

    def writeDfToSheet(self, sheetName, dfToWrite, startRow, startMarker, endMarker):
        ws = self.wb[sheetName]
        curr_row = startRow
        
        # Write start marker
        cell = ws.cell(curr_row, 1, startMarker)
        cell.font = Font(bold = True)
        curr_row += 1
        
        # Write headers
        for c_idx, col_name in enumerate(dfToWrite.columns, start=1):
            cell = ws.cell(curr_row, c_idx, col_name)
            cell.font = Font(bold=True)
        curr_row += 1
        
        # Write data rows
        for row_data in dfToWrite.itertuples(index=False):
            for c_idx, value in enumerate(row_data, start=1):
                #To ensure right formatting is applied when writing to Excel
                if isinstance(value, list):
                    cell_value = str(value)
                elif value in ['True', 'False']:
                    cell_value = True if value == 'True' else False
                else:
                    cell_value = value
                cell = ws.cell(curr_row, c_idx, cell_value)
            curr_row += 1
        
        # Write end marker
        cell = ws.cell(curr_row, 1, endMarker)
        cell.font = Font(bold=True)
        
        return curr_row + 1

    def rowsFinder(self, ws, startMarker, endMarker):
        start_row = None
        end_row = None
        
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == startMarker:
                start_row = idx
            if row[0] == endMarker:
                end_row = idx
        return start_row, end_row


    def excelToDfConverter(self, sheetName, startMarker, endMarker):
        ws = self.wb[sheetName]
        #Find range
        start_row, end_row = self.rowsFinder(ws, startMarker, endMarker)
        
        # Extract range
        data = list(ws.iter_rows(min_row=start_row+1, max_row=end_row-1, values_only=True))
        
        # First row is header
        headers = data[0]
        rows = data[1:]
        
        df = pd.DataFrame(rows, columns=headers)

        df.dropna(axis=1, how='all', inplace=True)
        
        # Sort by step number (assuming column is named 'step' or 'Step')
        if 'step' in df.columns:
            df = df.sort_values('step').reset_index(drop=True)
        return end_row, df

    def deleteRange(self, sheetName, startMarker, endMarker):
        ws = self.wb[sheetName]
        start_row, end_row = self.rowsFinder(ws, startMarker, endMarker)
        
        if start_row is None or end_row is None:
            return
        num_rows_to_delete = end_row - start_row + 1
        ws.delete_rows(start_row, num_rows_to_delete)

    def save_wb(self):
        #Delete an empty Sheet named "Sheet" if one exists
        try:
            self.wb.remove("Sheet")
        except:
            pass

        self.wb.save(self.filepath)
    
class CsvManager:
    @staticmethod
    def writeDfToCsv(df:pd.DataFrame, filepath:str):
        try:
            df.to_csv(filepath, index=False)
        except Exception as e:
            print(f'Unable to write to a csv: {e}')

    @staticmethod
    def readCsvToDf(filepath):
        try:
            df = pd.read_csv(filepath)
            return df
        except Exception as e:
            print(f'Unable to read from the csv {e}')